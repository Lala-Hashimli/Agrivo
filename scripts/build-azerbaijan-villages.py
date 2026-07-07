# -*- coding: utf-8 -*-
"""Build frontend/src/data/azerbaijanVillages.json from azerbaycan_kendleri_B_D_sutunlari.xlsx."""

from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "frontend/src/data/azerbaijanVillages.json"
DEFAULT_XLSX = Path.home() / "Downloads" / "azerbaycan_kendleri_B_D_sutunlari.xlsx"
FALLBACK_XLSX = Path.home() / "Downloads" / "azerbaycan_kendleri_B_D_sutunlari(1).xlsx"


def resolve_xlsx(path: Path | None = None) -> Path:
    if path and path.exists():
        return path
    if DEFAULT_XLSX.exists():
        return DEFAULT_XLSX
    if FALLBACK_XLSX.exists():
        return FALLBACK_XLSX
    raise FileNotFoundError(
        f"Excel file not found. Expected {DEFAULT_XLSX} or {FALLBACK_XLSX}",
    )


def main(xlsx_path: Path | None = None) -> None:
    source = resolve_xlsx(xlsx_path)
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheet_name = next((n for n in wb.sheetnames if "D" in n or "K" in n), wb.sheetnames[0])
    ws = wb[sheet_name]

    records: list[dict[str, str]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        district = str(row[0]).strip() if row[0] else ""
        village = str(row[1]).strip() if row[1] else ""
        if district and village:
            records.append({"district": district, "village": village})

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        json.dumps(records, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(f"Wrote {len(records)} village rows from {source.name} to {OUT}")
    wb.close()


if __name__ == "__main__":
    arg = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    main(arg)
